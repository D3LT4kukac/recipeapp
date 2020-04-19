from openpyxl import *
from tkinter import *
import openpyxl, os, re

#global variables for now... (its better to not have globals)
cwd = os.getcwd() #get current working directory
filepath = cwd + "/recipe_book.xlsx" # set filepath to the excel file


#begin TK GUI interface
root = Tk()

#definitions
def deleteChosen():
	#lista.delete(ACTIVE)
	book = load_workbook(filepath)
	ws = book.worksheets[0]
	listainfo = lista.get(ACTIVE)
	my_regex = r"\b(?=\w)" + re.escape(listainfo) + r"\b(?!\w)"
	for cell in ws["A"]:
		if re.search(my_regex, cell.value):
			ws.delete_rows(cell.row)
			lista.delete(ACTIVE)	
	book.save(filepath)

def openExcel():
	#os.system(filepath)
	os.system("start " + filepath)

global labelInList
labelInList = Label(root, text="Recipe name taken!\nPlease chooseanother name\nand try again!", fg="red")
#define labelerror
labelError = Label(root)
#saves entry data to excel file
def clickB():
	if nameE.get() in lista.get(0, END):
		
		labelInList.place(x = 390 , y = 100, anchor=NW)

	else:
		try:
			lista.insert("end", nameE.get())

			book = load_workbook(filepath)
			ws = book.worksheets[0]
			for cell in ws["A"]:
				if cell.value is None:
					ws.delete_rows(cell.row) #delete excel row if empty
					#print(cell.row)
					emptyRow = cell.row
					#break
			else:
				#print(cell.row + 1)
				emptyRow = cell.row + 1

			#print("first empty row is ", emptyRow)

			data = [(nameE.get(), recipeE.get("1.0", END))]
			for row in data:
				ws.append(row)

			book.save(filepath)


			#globalize errorlabel


			global labelError
			#clear entryboxes
			nameE.delete(0, END)
			recipeE.delete("1.0", END)
			
			labelError.destroy()
			labelInList.destroy()
			
		
		except IOError:
			#errormessage for open excel 
			labelError = Label(root, text="Something went wrong,\nPlease close Excel \nand try again",  fg="red")
			labelError.place(x = 400, y = 100, anchor = NW) 


#GUI settings
root.title("Recipe adding device,1.0")
root.geometry("550x300")

#add a listbox to GUI
lista = Listbox(root)
lista.place(x = 2, y = 70, anchor = NW, width=185, height= 200)
 
#lista label
listalabel = Label(root, text="List of recipes:")
listalabel.place(x=2, y=48)

#GUI oopen excel button
openB = Button(root, text="Open Excel file", bg="white", fg="purple", command=openExcel)
openB.place(x = 240, y = 272, anchor = NW) 

#GUI delete data
deleteB = Button(root, text="Remove selected item", bg="red", command=deleteChosen)
deleteB.place(x = 29, y = 272, anchor = NW)

#GUI add recipe button
addB = Button(root, text="Add new recipe\nto the list", fg="white", bg="green", command=clickB)
addB.place(x = 420, y = 2, anchor = NW)

#GUI add recipe name entry box
nameE = Entry(root, width=30)
nameE.place(x = 190, y = 2, anchor = NW)

#GUI scrollbar for recipeE
scrolly = Scrollbar(root)

recipeE = Text(root, yscrollcommand=scrolly.set, width=23, height=15)

scrolly.config(command=recipeE.yview)
scrolly.place(x = 380, y = 25, anchor = NW, height=245) 

#GUI add recipe description box
#recipeE = Text(root, yscrollcommand=scrolly.set, width=20)
recipeE.place(x = 190, y = 25, anchor = NW) 
#GUI label for name of recipe
labelN = Label(root, text="Enter the name of the new recipe:", bg="white")
labelN.place(x = 5, y = 2, anchor = NW)

#GUI label for description of recipe
labelD = Label(root, text="Enter the description of the recipe:", bg="white")
labelD.place(x = 2, y = 25, anchor = NW)



#read the excel file  
try:
	wb = openpyxl.load_workbook(filepath) 
except:
	wb = Workbook()
	sheet = wb["Sheet"]

#just to make sure its there	
sheet = wb["Sheet"] 
#excel max.row finding
m = sheet.max_row 

#for i in range(start of i, end of i, stepping of i [if not set then i moves up by 1 increment])
#adds excel column data to list in GUI
for i in range(1, m + 1): 
    cell = sheet.cell(row = i, column = 1) 
    lista.insert("end", cell.value)
    print(cell.value) 
    
wb.save("recipe_book.xlsx")

root.mainloop()